from openpyxl.styles import PatternFill, Font


GREEN = PatternFill(fill_type="solid", start_color="90EE90")
YELLOW = PatternFill(fill_type="solid", start_color="FFF799")
RED = PatternFill(fill_type="solid", start_color="FFC7CE")
GOLD = PatternFill(fill_type="solid", start_color="FFD966")


def format_sheet(ws):

    # Bold Header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Highlight Benchmark Row
    benchmark_col = None

    for cell in ws[1]:
        if cell.value == "is_benchmark":
            benchmark_col = cell.column
            break

    if benchmark_col:

        for row in range(2, ws.max_row + 1):

            if ws.cell(row, benchmark_col).value == "1":

                for c in range(1, ws.max_column + 1):
                    ws.cell(row, c).fill = GOLD

    # Color Percentiles
    for col in ws.iter_cols():

        header = col[0].value

        if header == "percentile_rank":

            for cell in col[1:]:

                if cell.value is None:
                    continue

                if cell.value >= 75:
                    cell.fill = GREEN

                elif cell.value >= 25:
                    cell.fill = YELLOW

                else:
                    cell.fill = RED