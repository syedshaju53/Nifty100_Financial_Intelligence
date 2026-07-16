import sqlite3
import pandas as pd
from pathlib import Path

from openpyxl import load_workbook

from src.analytics.peer_report import format_sheet

BASE_DIR = Path(__file__).resolve().parents[2]

DB_PATH = BASE_DIR / "db" / "nifty100.db"

OUTPUT = BASE_DIR / "output"

OUTPUT.mkdir(exist_ok=True)

conn = sqlite3.connect(DB_PATH)

master = pd.read_sql(
    "SELECT * FROM master_financials",
    conn
)

peer = pd.read_sql(
    "SELECT * FROM peer_groups",
    conn
)

percentiles = pd.read_sql(
    "SELECT * FROM peer_percentiles",
    conn
)

merged = (
    peer
    .merge(master, on="company_id", how="left")
    .merge(
    percentiles[
        [
            "company_id",
            "metric",
            "percentile_rank",
            "year"
        ]
    ],
    on=["company_id", "year"],
    how="left"
)
)

excel_path = OUTPUT / "peer_comparison.xlsx"

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:

    groups = merged["peer_group_name"].dropna().unique()

    print()

    print("Peer Groups :", len(groups))

    for group in groups:

        df = merged[
            merged["peer_group_name"] == group
        ].copy()

        numeric = df.select_dtypes(include="number")

        median_row = numeric.median()

        median = pd.DataFrame(
            [median_row],
            columns=numeric.columns
        )

        median["company_id"] = "PEER_MEDIAN"

        df = pd.concat(
            [df, median],
            ignore_index=True
        )

        df.to_excel(
            writer,
            sheet_name=group[:31],
            index=False
        )

workbook = load_workbook(excel_path)

for sheet in workbook.sheetnames:

    format_sheet(workbook[sheet])

workbook.save(excel_path)

print()

print("Peer Comparison Report Generated!")

print(excel_path)

conn.close()